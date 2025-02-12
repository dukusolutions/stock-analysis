import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def calculate_technical_indicators(stock_data):
    """
    Calculate technical indicators for trend analysis
    """
    df = stock_data.copy()
    
    # Calculate 50-day and 200-day moving averages
    df['MA50'] = df['Close'].rolling(window=50).mean()
    df['MA200'] = df['Close'].rolling(window=200).mean()
    
    # Calculate RSI (14-day)
    delta = df['Close'].diff()
    gain = (delta.where(delta > 0, 0)).rolling(window=14).mean()
    loss = (-delta.where(delta < 0, 0)).rolling(window=14).mean()
    rs = gain / loss
    df['RSI'] = 100 - (100 / (1 + rs))
    
    # Calculate MACD
    exp1 = df['Close'].ewm(span=12, adjust=False).mean()
    exp2 = df['Close'].ewm(span=26, adjust=False).mean()
    df['MACD'] = exp1 - exp2
    df['Signal_Line'] = df['MACD'].ewm(span=9, adjust=False).mean()
    
    # Handle any NaN values
    for col in ['MA50', 'MA200', 'RSI', 'MACD', 'Signal_Line']:

        df[col] = df[col].fillna(0)
    
    return df

def identify_value_pattern(stock_data):
    """
    Identify potential value opportunities
    Returns a dictionary of pattern indicators and their values
    """
    try:
        # Extract scalar values using .iloc[0] for Series
        current_price = stock_data['Close'].iloc[-1]
        high_price = stock_data['High'].max()
        low_price = stock_data['Low'].min()
        avg_volume = stock_data['Volume'].mean()
        current_volume = stock_data['Volume'].iloc[-1]
        current_rsi = stock_data['RSI'].iloc[-1]
        current_ma200 = stock_data['MA200'].iloc[-1]
        volatility = stock_data['Close'].pct_change().std()
        
        # Calculate base metrics
        pct_from_high = ((high_price - current_price) / high_price * 100)
        volume_ratio = current_volume / avg_volume if avg_volume > 0 else 1.0
        
        # Pattern checks
        pattern = {
            'Deep_Value': pct_from_high > 25,
            'High_Volume': volume_ratio > 1.2,
            'RSI_Oversold': current_rsi < 35,
            'Near_Support': current_price <= low_price * 1.05,
            'Below_MA200': current_price < current_ma200,
            'Stabilizing': volatility < 0.02
        }
        
        # Calculate pattern score
        pattern_score = sum([
            30 if pattern['Deep_Value'] else 0,
            20 if pattern['Near_Support'] else 0,
            15 if pattern['High_Volume'] else 0,
            15 if pattern['RSI_Oversold'] else 0,
            10 if pattern['Below_MA200'] else 0,
            10 if pattern['Stabilizing'] else 0
        ])
        
        pattern['Pattern_Score'] = float(pattern_score)
        pattern['Pattern_Detected'] = pattern_score >= 60
        
        return pattern
        
    except Exception as e:
        print(f"Warning in pattern identification: {str(e)}")
        return {
            'Deep_Value': False,
            'Near_Support': False,
            'High_Volume': False,
            'RSI_Oversold': False,
            'Below_MA200': False,
            'Stabilizing': False,
            'Pattern_Score': 0.0,
            'Pattern_Detected': False
        }

def get_fundamental_metrics(ticker):
    """
    Fetch fundamental metrics from yfinance
    """
    try:
        stock = yf.Ticker(ticker)
        info = stock.info
        
        return {
            'Market_Cap': info.get('marketCap'),
            'PE_Ratio': info.get('forwardPE'),
            'PEG_Ratio': info.get('pegRatio'),
            'Price_to_Book': info.get('priceToBook'),
            'Debt_to_Equity': info.get('debtToEquity'),
            'Current_Ratio': info.get('currentRatio'),
            'Profit_Margins': info.get('profitMargins'),
            'ROE': info.get('returnOnEquity'),
            'Dividend_Rate': info.get('dividendRate'),
            'Dividend_Yield': info.get('dividendYield'),
            'Beta': info.get('beta'),
            'Sector': info.get('sector'),
            'Industry': info.get('industry')
        }
    except Exception as e:
        print(f"Error fetching fundamental data for {ticker}: {str(e)}")
        return {}

def calculate_enhanced_value_score(metrics, pattern_data):
    """
    Calculate an enhanced value score incorporating both fundamental and technical factors
    """
    score = 50  # Base score
    
    # Pattern recognition score (0-30 points)
    pattern_score = pattern_data.get('Pattern_Score', 0)
    score += pattern_score * 0.3
    
    # Traditional value metrics (0-40 points)
    if metrics.get('PE_Ratio'):
        if metrics['PE_Ratio'] < 15:
            score += 10
        elif metrics['PE_Ratio'] < 20:
            score += 5
            
    if metrics.get('Dividend_Yield'):
        div_yield = float(metrics['Dividend_Yield']) if metrics['Dividend_Yield'] else 0
        if div_yield > 0.04:  # 4% yield
            score += 10
        elif div_yield > 0.02:  # 2% yield
            score += 5
            
    if metrics.get('Price_to_Book'):
        if metrics['Price_to_Book'] < 1.5:
            score += 10
        elif metrics['Price_to_Book'] < 3:
            score += 5
    
    # Market position and stability (0-30 points)
    if metrics.get('Market_Cap'):
        if metrics['Market_Cap'] > 10e9:  # 10B+ market cap
            score += 10
            
    if metrics.get('Beta'):
        if metrics['Beta'] < 1.2:  # Lower volatility
            score += 10
            
    if metrics.get('Current_Ratio'):
        if metrics['Current_Ratio'] > 1.5:
            score += 10
            
    return min(max(score, 0), 100)  # Ensure score is between 0 and 100

def analyze_stocks(tickers=None, lookback_days=365):
    """
    Comprehensive stock analysis combining technical and fundamental factors
    """
    if tickers is None:
        print("Fetching S&P 500 tickers...")
        sp500 = pd.read_html('https://en.wikipedia.org/wiki/List_of_S%26P_500_companies')[0]
        tickers = sp500['Symbol'].tolist()
    
    end_date = datetime.now()
    start_date = end_date - timedelta(days=lookback_days)
    print(f"\nFetching data from {start_date.date()} to {end_date.date()}")
    
    results = []
    
    for ticker in tickers:
        try:
            print(f"\nAnalyzing {ticker}...")
            
            # Create ticker object first
            ticker_obj = yf.Ticker(ticker)
            
            # Download historical data
            stock_data = yf.download(ticker, start=start_date, end=end_date, progress=False)
            if stock_data.empty:
                print(f"No data found for {ticker}")
                continue
                
            # Handle multi-level columns - flatten the structure
            stock_data.columns = [col[0] for col in stock_data.columns]
            
            print("\nColumn names after cleanup:", stock_data.columns.tolist())
            print("\nFirst few rows of cleaned data:")
            print(stock_data.head())
                
            # Calculate technical indicators
            stock_data = calculate_technical_indicators(stock_data)
            
            # Get pattern data
            pattern_data = identify_value_pattern(stock_data)
            
            # Get fundamental data
            fund_data = get_fundamental_metrics(ticker)
            
            # Get clean scalar values
            current_price = stock_data['Close'].iloc[-1]
            high_price = stock_data['High'].max()
            low_price = stock_data['Low'].min()
            
            print(f"Current price: {current_price}")
            print(f"High price: {high_price}")
            print(f"Low price: {low_price}")
            
            # Calculate 52-week high
            fifty_two_week_high = stock_data['High'].rolling(window=252, min_periods=1).max().iloc[-1]
            
            print(f"52-week high: {fifty_two_week_high}")
            
            # Calculate percentage from high
            if fifty_two_week_high > 0:
                pct_from_high = ((fifty_two_week_high - current_price) / fifty_two_week_high) * 100
                print(f"Calculated percentage from high: {pct_from_high:.2f}%")
            else:
                print("Invalid 52-week high value")
                pct_from_high = 0.0
            
            # Create result dictionary with scalar values
            result = {
                'Company': ticker_obj.info.get('shortName', ticker),  # Use get() with default value
                'Ticker': str(ticker),
                'Analysis_Date': datetime.now().strftime('%Y-%m-%d'),
                'Current_Price': float(current_price),
                'Value_Score': float(calculate_enhanced_value_score(fund_data, pattern_data)),
                'Pattern_Score': float(pattern_data['Pattern_Score']),
                'Pct_From_52W_High': float(pct_from_high),
                'Pattern_Detected': bool(pattern_data['Pattern_Detected'])
            }
            
            # Add pattern detection results
            for key in ['Deep_Value', 'Near_Support', 'High_Volume', 
                       'RSI_Oversold', 'Below_MA200', 'Stabilizing']:
                result[key] = bool(pattern_data.get(key, False))
            
            # Add fundamental metrics
            for key, value in fund_data.items():
                if isinstance(value, (int, float)):
                    result[key] = float(value) if value is not None else None
                else:
                    result[key] = value
            
            results.append(result)
            
        except Exception as e:
            print(f"Error analyzing {ticker}: {str(e)}")
            continue
    
    # Create DataFrame
    df = pd.DataFrame(results)
    
    # Calculate watchlist flags
    if not df.empty:
        df['Watchlist'] = (df['Pattern_Score'] > 60) & (df['Value_Score'] > 70)
        df['Priority_Watch'] = (df['Pattern_Score'] > 70) & (df['Value_Score'] > 80)
    
    # Add timestamp to filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    # Save to Excel
    save_to_enhanced_excel(df, f'stock_analysis_{timestamp}.xlsx')
    
    return df

def save_to_enhanced_excel(df, filename):
    """
    Save analysis results to Excel with enhanced formatting
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Analysis"
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    watchlist_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    priority_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    
    # Write headers
    headers = list(df.columns)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Write data with conditional formatting
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Apply watchlist highlighting
            if headers[col_idx-1] == 'Watchlist' and value:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=c).fill = watchlist_fill
            
            # Apply priority watch highlighting
            if headers[col_idx-1] == 'Priority_Watch' and value:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=c).fill = priority_fill
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[get_column_letter(column[0].column)].width = max_length + 2
    
    # Add filters
    ws.auto_filter.ref = ws.dimensions
    
    # Save workbook
    wb.save(filename)

if __name__ == "__main__":
    # Test with just two tickers for detailed debugging
    test_tickers = ['AAPL', 'MSFT']
    results = analyze_stocks()
    
    print("\nRaw DataFrame Contents:")
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', None)
    print(results)
    
    print("\nAnalysis Results:")
    if not results.empty:
        # Display relevant columns
        display_columns = ['Company', 'Ticker', 'Current_Price', 'Value_Score', 
                         'Pattern_Score', 'Pct_From_52W_High', 'Pattern_Detected']
        
        # Display all results
        print("\nAll analyzed stocks:")
        print(results[display_columns].to_string())
        
        # Display high value opportunities
        high_value = results[results['Value_Score'] > 70]
        if not high_value.empty:
            print("\nHigh Value Opportunities:")
            print(high_value[display_columns].to_string())
        
        # Display strong pattern matches
        strong_pattern = results[results['Pattern_Score'] > 60]
        if not strong_pattern.empty:
            print("\nStrong Pattern Matches:")
            print(strong_pattern[display_columns].to_string())
    else:
        print("No results were found. Please check the errors above.")